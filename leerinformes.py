import os
import openpyxl
from flask import Flask

app = Flask(__name__)
titulo = "Informes"

datos = []

for archivo in os.listdir('.'):
    if archivo.endswith("xlsx"):
        wb = openpyxl.load_workbook(archivo, data_only=True)
        main = wb.sheetnames[0]
        sheet = wb[main]
        #armo una lista en donde voy a incluir los diccionario
        listaDicc = []
        #inicio un loop para explorar las celdas. comienzo explorando las filas
        for x in range(1, sheet.max_row+1):
            #me fijo si la primera celda esta completa, así puedo saber que es fila de etiquetas
            if sheet.cell(row=x, column=1).value != None :
                #armo el key del diccionario que va a llevar el nombre de la fila
                k = sheet.cell(row=x, column=1).value
                #agrego a la lista el diccionario que tiene el key y asigna como valor un diccionario vacio
                listaDicc.append({k:{}})
                #inicio un loop para enumerar los datos de la fila de abajo de la fila que tiene las etiquetas
                for y in range(1, sheet.max_column+1):
                    #exploro si cada celda tiene un valor
                    if sheet.cell(row=x+1, column=y).value != None:
                        #si la celda tiene valor actualizo el diccionario con la etiqueta como key y el valor como value
                        listaDicc[-1][k].update({sheet.cell(row=x, column=y).value:str(sheet.cell(row=x+1, column=y).value)})
        datos.append(listaDicc)

datosHTML = ""
for i in datos:
    for index,item in enumerate(i):
        for key,value, in item.items():
            if len(value)>0:
                datosHTML += "<div><h1>"+key+"</h1>"
                for k,v in value.items():
                    datosHTML += "<h2>"+k+": "+v+"<h2>" 
                datosHTML += "</div>"

@app.route("/")
def home_www():
    return (f"""<title>{titulo}</title> {datosHTML}""")

app.run(host="localhost", port=8080, debug=True)