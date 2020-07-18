import os
import openpyxl

wb = openpyxl.load_workbook("Gonzales Ivan Guido 0620 BD.xlsx")
#wb = openpyxl.load_workbook("Gonzales Ivan Guido 0620 BD.xlsx", data_only=True)
main = wb.sheetnames[0]
sheet = wb[main]
celda = sheet["B2"]

#for i in range(1,15):
#    print(sheet.cell(column=i,row=2).value)

#print(sheet.max_column)
#print(sheet.max_row)

for i in range(1,sheet.max_row):
    if sheet.cell(column=sheet.max_column,row=i).value != None:
        print(sheet.cell(column=sheet.max_column,row=i).value)